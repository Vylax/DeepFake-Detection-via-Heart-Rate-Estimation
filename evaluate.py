import sys
import numpy as np

sys.path.append('./cohface')
import cf_hr
import em_hr

import openpyxl
import os

import inject_hr

#Set these variables
dataset_is_cohface = False #True if the dataset is cohface, False otherwise
db_path = '../ff++/' #This should be the path to the dataset folder, it should have the protocols and videos inside it or in a subfolder
file_ext = '.mp4' #This should be the extension of the videos in the dataset
protocol = 'protocols/all.txt' #This should be the relative path of the protocol you want to use (the protocol is a file listing the relative path of all videos you want to evaluate without their extension)
injected_heartrates = [65, 80, 90, 100] #This should be a list of the heart rates you want to inject and evaluate, if you don't want to inject any hearrate just set it to [-1]

#Don't set these variables
gtdatapath = ''
estimations = []
files_paths = []

def eval_data(data_path):
    ground_truth_hr = cf_hr.get_heart_rate(gtdatapath) if dataset_is_cohface else -333
    eulerian_magnification_hr = em_hr.get_heart_rate(data_path+file_ext)
    
    return ground_truth_hr,eulerian_magnification_hr

def find_next_filename(base_filename):
    i = 1
    while os.path.exists(f"{base_filename}{i}.xlsx"):
        i += 1
    return f"{base_filename}{i}.xlsx"

def write_excel_file():
    # Create a new Excel workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set headers
    if injected_heartrates == [-1]:
        headers = ["sample", "gt", "est"]
    else:
        headers = ["sample", "gt", "est"] + [f"est{rate}" for rate in injected_heartrates]

    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

    # Fill data
    for row, estimation in enumerate(estimations, start=2):
        for col, value in enumerate(estimation[:3], start=1):
            sheet.cell(row=row, column=col, value=value)
        if injected_heartrates != [-1]:
            for i, rate in enumerate(injected_heartrates):
                est_index = 3 + i + 1
                sheet.cell(row=row, column=est_index, value=estimation[3 + i])

    # Save the workbook to a file
    workbook.save(find_next_filename("output"))

with open(db_path+protocol, 'r') as file:
    for line in file:
        # Process each line as needed
        files_paths.append(db_path+line.strip())

#Example execution
for hr in injected_heartrates:
    # Iterate over each line in the file
    i = -1
    for file_path in files_paths:
        i+=1
        try:
            if file_path[-1:] == "a": #This is a very unclean way to do this but it works fine...
                gtdatapath=file_path+'.hdf5'
            
            #create injected video
            if(hr!=-1):
                inject_hr.inject_heart_rate(file_path+file_ext, file_path+'_hacked___'+str(hr)+file_ext, hr)
            
            if(hr==injected_heartrates[0]):
                gt,est=eval_data(file_path)
            gth,esth=eval_data(file_path+'_hacked___'+str(hr))
        except Exception as e:
            print(f"Error with sample: {file_path}\nError: {str(e)}")
        else:
            if(hr==injected_heartrates[0]):
                estimations.append([file_path,gt,est,esth])
            else:
                estimations[i].append(esth)
            print(file_path,estimations[i][2],hr,esth)

write_excel_file()