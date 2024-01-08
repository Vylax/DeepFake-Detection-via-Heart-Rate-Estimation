import sys
import numpy as np
#sys.path.append('../cohface/cohface')

sys.path.append('./cohface')
import cf_hr
import em_hr

import openpyxl
import os

import inject_hr

gtdatapath = ''

def eval_data(data_path):
    #data_path should be equal to the path of the folder containing data.avi and data.hdf5
    
    ground_truth_hr = 123#cf_hr.get_heart_rate(gtdatapath)
    eulerian_magnification_hr = em_hr.get_heart_rate(data_path+'.mp4')
    
    return ground_truth_hr,eulerian_magnification_hr

db_path = '../ff++/'

estimations = []

target_hr = 65

files_paths = []

with open(db_path+'protocols/all.txt', 'r') as file:
    for line in file:
        # Process each line as needed
        files_paths.append(db_path+line.strip()) 

hue = 1.7222
for hr in [90, 65, 120]:#list(np.linspace(1.5, 2.5, 10))

    #print(hue)
    # Iterate over each line in the file
    for file_path in files_paths:
        try:
            if file_path[-1:] == "a":
                gtdatapath=file_path+'.hdf5'
            
            #create injected video
            inject_hr.inject_heart_rate(file_path+'.mp4', file_path+'_hacked'+str(hr)+'.mp4', hr, hue)
            
            #print("gtdatapath",gtdatapath,'|'+file_path[-1:]+'|')
            #gt,est=eval_data(file_path)
            #TODO remove when done skipping original videos
            gt,est=123,456
            gth,esth=eval_data(file_path+'_hacked'+str(hr))
            gth=hr
            diff=gt-est
        except Exception as e:
            print(f"Error with sample: {file_path}\nError: {str(e)}")
        else:
            print(file_path,hr,esth,est)
            #estimations.append([file_path,esth])
            #hues[index][1].append(esth)

def find_next_filename(base_filename):
    i = 1
    while os.path.exists(f"{base_filename}{i}.xlsx"):
        i += 1
    return f"{base_filename}{i}.xlsx"

def write_excel_file():
    # Create a new Excel workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(row=1, column=1, value="sample")
    sheet.cell(row=1, column=2, value="est")
    for i in range(len(estimations)):
        sheet.cell(row=i+2, column=1, value=estimations[i][0])
        sheet.cell(row=i+2, column=2, value=estimations[i][1])

    # Save the workbook to a file
    workbook.save(find_next_filename("output"))

#print(hues)

#write_excel_file()