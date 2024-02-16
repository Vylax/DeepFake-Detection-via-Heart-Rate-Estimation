import sys
import numpy as np
#sys.path.append('../cohface/cohface')

sys.path.append('./cohface')
import cf_hr
import em_hr

import openpyxl
import os

import inject_hr3
import time

start_time = time.time()

def eval_data(data_path):
    eulerian_magnification_hr = em_hr.get_heart_rate(data_path+file_ext)
    
    return 333,eulerian_magnification_hr

db_path = '../'

estimations = []

files_paths = []

file_ext = '.mp4'

with open(db_path+'selection.txt', 'r') as file:
    for line in file:
        # Process each line as needed
        files_paths.append(db_path+line.strip()) 

hue = 1.7222
for hr in [65, 80, 90, 100]:
    # Iterate over each line in the file
    i = -1
    for file_path in files_paths:
        i+=1
        try:
            #create injected video
            inject_hr3.inject_heart_rate(file_path+file_ext, file_path+'_hacked___'+str(hr)+file_ext, hr, hue)
            
            if(hr==65):
                gt,est=eval_data(file_path)
            gth,esth=eval_data(file_path+'_hacked___'+str(hr))
        except Exception as e:
            print(f"Error with sample: {file_path}\nError: {str(e)}")
            print("Time: "+str(time.time()-start_time)+"s")
        else:
            if(hr==65):
                estimations.append([file_path,gt,est,esth])
            else:
                estimations[i].append(esth)
            print(file_path,estimations[i][2],hr,esth)
            print("Time: "+str(time.time()-start_time)+"s")

def find_next_filename(base_filename):
    i = 1
    while os.path.exists(f"{base_filename}{i}.xlsx"):
        i += 1
    return f"{base_filename}{i}.xlsx"

def write_excel_file():
    # Create a new Excel workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(row=1, column=1, value="sample")
    sheet.cell(row=1, column=2, value="gt")
    sheet.cell(row=1, column=3, value="est")
    sheet.cell(row=1, column=4, value="est65")
    sheet.cell(row=1, column=5, value="est80")
    sheet.cell(row=1, column=6, value="est90")
    sheet.cell(row=1, column=7, value="est100")
    for i in range(len(estimations)):
        sheet.cell(row=i+2, column=1, value=estimations[i][0])
        sheet.cell(row=i+2, column=2, value=estimations[i][1])
        sheet.cell(row=i+2, column=3, value=estimations[i][2])
        sheet.cell(row=i+2, column=4, value=estimations[i][3])
        sheet.cell(row=i+2, column=5, value=estimations[i][4])
        sheet.cell(row=i+2, column=6, value=estimations[i][5])
        sheet.cell(row=i+2, column=7, value=estimations[i][6])

    # Save the workbook to a file
    workbook.save(find_next_filename("output_eval_celebdf"))

write_excel_file()